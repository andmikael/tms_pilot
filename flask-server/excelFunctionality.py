from flask import Blueprint, jsonify, request
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from flask_cors import cross_origin
from datetime import date

# Create a Blueprint object for Excel-related functions
excel_bp = Blueprint('excel', __name__)


"""
Handles the upload of route data and saves it into a new Excel file.

:param route_name: str, name of the route to be saved
:param file_name: str, original file name for the route data
:param excel_data: list, list of dictionaries containing route data
:param start_location: dict, dictionary with the start location details
:param end_location: dict, dictionary with the end location details
:return: JSON response, success or failure message
"""
@excel_bp.route('/api/upload', methods=['POST'])
@cross_origin()
def upload_excel():
    # Parse JSON data from the request body
    data = request.get_json()
    # Retrieve route name, file name, Excel data, start, and end locations from the request.
    route_name = data.get("routeName", "Uusi_reitti")
    file_name = data.get("fileName", "Tuntematon_tiedostonimi")
    excel_data = data.get("data", [])
    start_location = data.get("startLocation", {})
    end_location = data.get("endLocation", {})

    # Create a new Excel workbook and set the active worksheet title.
    wb = Workbook()
    ws = wb.active
    ws.title = "Luetut paikat"

    # Determine the save directory and create it if it does not exist.
    save_dir = os.path.join(".secret", "ExcelFiles")
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)

    # Remove extension from file_name for reuse in constructing the final file name.
    original_file_name = os.path.splitext(file_name)[0]
    file_path = os.path.join(save_dir, f"{route_name} ({original_file_name}).xlsx")
    file_counter = 1
    # If a file with this name exists, increment the counter until a unique file name is created.
    if os.path.exists(file_path):
        while os.path.exists(file_path):
            file_path = os.path.join(save_dir, f"{route_name} {file_counter} ({original_file_name}).xlsx")
            file_counter += 1
    
    today = date.today()
    # Construct final file name to be written into the Excel file.
    final_file_name = os.path.splitext(os.path.basename(file_path))[0]
    # Write the route name in cell B1 with a label in A1.
    ws["A1"] = "Reitin nimi:"
    ws["B1"] = final_file_name
    ws["C1"] = original_file_name
    ws["D1"] = today.strftime("%d.%m.%Y")

    # Set up header for the Excel sheet.
    header = ["Nimi", "Osoite", "Postinumero", "Kaupunki", "Vakionouto", "Lat", "Lon"]
    ws.append(header)

    # If start_location data exists, write it to predefined cells starting from I1.
    if start_location:
        ws["I1"] = start_location.get("name")
        ws["J1"] = start_location.get("address")
        ws["K1"] = start_location.get("postalCode")
        ws["L1"] = start_location.get("city")
        ws["M1"] = "yes"
        ws["N1"] = start_location.get("departureTime")
        ws["O1"] = start_location.get("lat")
        ws["P1"] = start_location.get("lon")

    # If end_location data exists, write it to predefined cells starting from I2.
    if end_location:
        ws["I2"] = end_location.get("name")
        ws["J2"] = end_location.get("address")
        ws["K2"] = end_location.get("postalCode")
        ws["L2"] = end_location.get("city")
        ws["M2"] = "yes"
        ws["N2"] = end_location.get("endTime")
        ws["O2"] = end_location.get("lat")
        ws["P2"] = end_location.get("lon")

    # Write each additional route point starting from row 3.
    row_num = 3
    for item in excel_data:
        ws.cell(row=row_num, column=1, value=item.get("name"))
        ws.cell(row=row_num, column=2, value=item.get("address"))
        ws.cell(row=row_num, column=3, value=item.get("postalCode"))
        ws.cell(row=row_num, column=4, value=item.get("city"))
        ws.cell(row=row_num, column=5, value=item.get("standardPickup"))
        ws.cell(row=row_num, column=6, value=item.get("lat"))
        ws.cell(row=row_num, column=7, value=item.get("lon"))
        row_num += 1

    try:
        # Save the workbook to the specified file path.
        wb.save(file_path)
        return jsonify({"message": "Tiedosto tallennettu onnistuneesti", "file_path": file_path}), 200
    except Exception as e:
        # Return an error response if saving fails.
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}), 500

"""
Retrieves the list of all Excel files stored in the server directory.

:return: JSON response, dictionary containing filenames and paths of all .xlsx files
"""
@excel_bp.route('/api/get_excel_files', methods=['GET'])
@cross_origin()
def get_excel_jsons():
    # Define the folder where Excel files are stored.
    folder = os.path.join(".secret", "ExcelFiles")
    excel_jsons = {}
    
    if os.path.exists(folder):
        # List and process every Excel file in the folder.
        for file_name in os.listdir(folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(folder, file_name)
                try:
                    # Get the base name of the file (without extension).
                    base_name = os.path.splitext(file_name)[0]
                    excel_jsons[base_name] = {"file_name": file_name, "path": file_path}
                except Exception as e:
                    print(f"Error reading file {file_name}: {e}")
                    base_name = os.path.splitext(file_name)[0]
                    excel_jsons[base_name] = {"error": str(e)}
    else:
        # If the folder does not exist, return a 404 error.
        return jsonify({"error": "Kansiota ei löydy"}), 404
    
    # Return the JSON object containing file data.
    return jsonify(excel_jsons)

"""
Deletes a specified Excel file from the server.

:param file_name: str, the name of the Excel file to be deleted
:return: JSON response, success or failure message
"""
@excel_bp.route('/api/delete_excel', methods=['DELETE'])
@cross_origin()
def delete_excel():
    data = request.get_json()
    file_name = data.get("file_name")
    # Check if file_name parameter is provided.
    if not file_name:
        return jsonify({"error": "file_name not provided"}), 400

    # Ensure the file name ends with ".xlsx"
    if not file_name.endswith(".xlsx"):
        file_name_full = file_name + ".xlsx"
    else:
        file_name_full = file_name

    folder = os.path.join(".secret", "ExcelFiles")
    file_path = os.path.join(folder, file_name_full)
    
    # Return error if file does not exist.
    if not os.path.exists(file_path):
        return jsonify({"error": "Tiedostoa ei löytynyt"}), 404

    try:
        # Attempt to remove the file.
        os.remove(file_path)
        return jsonify({"message": "Tiedosto poistettu onnistuneesti", "file_name": file_name_full}), 200
    except Exception as e:
        # Return error if deletion fails.
        return jsonify({"error": True, "message": f"Tiedoston poisto epäonnistui: {e}"}), 500

"""
Reads and returns the content of a specified Excel file, including route details.

:return: JSON response, dictionary containing route information from the Excel file
"""
@excel_bp.route('/api/get_route', methods=['GET'])
@cross_origin()
def get_excel_routes():
    folder = os.path.join(".secret", "ExcelFiles")
    excel_routes = {}
    
    # Check if the designated folder exists.
    if not os.path.exists(folder):
        return jsonify({"error": "Kansiota ei löydy"}), 404
    
    # Iterate over each Excel file in the folder.
    for file_name in os.listdir(folder):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder, file_name)
            try:
                # Load the workbook and select the active worksheet.
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                # Read the route name from cell B1.
                route_name = ws["B1"].value

                # Read the start location from cells I1 to P1.
                startPlace = {
                    "name": ws["I1"].value,
                    "address": ws["J1"].value,
                    "postalCode": ws["K1"].value,
                    "city": ws["L1"].value,
                    "standardPickup": ws["M1"].value,
                    "lat": ws["O1"].value,
                    "lon": ws["P1"].value,
                }
                
                # Read the end location from cells I2 to P2.
                endPlace = {
                    "name": ws["I2"].value,
                    "address": ws["J2"].value,
                    "postalCode": ws["K2"].value,
                    "city": ws["L2"].value,
                    "standardPickup": ws["M2"].value,
                    "lat": ws["O2"].value,
                    "lon": ws["P2"].value,
                }
                
                # Get start and end times from cells N1 and N2.
                startTime = ws["N1"].value
                endTime = ws["N2"].value

                # Read all route entries starting from row 3.
                routes = []
                # It is assumed that columns A to G hold the route data.
                for row in ws.iter_rows(min_row=3, max_col=7):
                    # Skip the row if it is completely empty.
                    if not any(cell.value for cell in row):
                        continue
                    route_entry = {
                        "name": row[0].value,        # Column A
                        "address": row[1].value,     # Column B
                        "postalCode": row[2].value,  # Column C
                        "city": row[3].value,        # Column D
                        "standardPickup": row[4].value,  # Column E
                        "lat": row[5].value,         # Column F
                        "lon": row[6].value,         # Column G
                    }
                    routes.append(route_entry)
                
                # Construct a JSON object for the route.
                route_json = {
                    "name": route_name,
                    "startPlace": startPlace,
                    "endPlace": endPlace,
                    "startTime": startTime,
                    "endTime": endTime,
                    "routes": routes,
                }
                # Add this route info using the file name as the key.
                excel_routes[file_name] = route_json
                
            except Exception as e:
                # In case of errors, add an error message for the specific file.
                excel_routes[file_name] = {"error": str(e)}
    
    # Return all the parsed routes in JSON format.
    return jsonify(excel_routes)

"""
Appends a new pickup point to an existing Excel file.

:param filename: str, the name of the Excel file to be updated
:param pickup: dict, dictionary containing pickup point details
:return: JSON response, success or failure message
"""
@excel_bp.route('/api/append_to_excel', methods=['POST'])
@cross_origin()
def append_to_excel():
    data = request.get_json()
    filename = data.get("filename")
    pickup = data.get("data")

    # Validate that all required fields are present in the pickup data.
    required_fields = ["name", "address", "postalCode", "city", "standardPickup", "lat", "lon"]
    if not filename or not all(field in pickup for field in required_fields):
        return jsonify({"error": "Puuttuvia kenttiä datasta."}), 400

    # Construct the file path of the Excel file.
    file_path = os.path.join(".secret", "ExcelFiles", f"{filename}.xlsx")
    if not os.path.exists(file_path):
        return jsonify({"error": "Excel-tiedostoa ei löydy."}), 404

    # Load the workbook and select the active worksheet.
    wb = load_workbook(file_path)
    ws = wb.active

    # Append the new pickup row to the worksheet.
    ws.append([
        pickup["name"],
        pickup["address"],
        pickup["postalCode"],
        pickup["city"],
        pickup["standardPickup"],
        pickup["lat"],
        pickup["lon"]
    ])

    try:
        # Save the updated workbook.
        wb.save(file_path)
        return jsonify({"message": "Paikka lisätty onnistuneesti."}), 200
    except Exception as e:
        # Return error message in case saving fails.
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}), 500

"""
Removes a specified pickup point from an Excel file.

:param filename: str, the name of the Excel file to be modified
:param pickup: dict, dictionary containing pickup point details to be removed
:return: JSON response, success or failure message
"""
@excel_bp.route('/api/remove_from_excel', methods=['POST'])
@cross_origin()
def remove_from_excel():
    data = request.get_json()
    filename = data.get("filename")
    pickup = data.get("data")

    # Validate required fields in the pickup data.
    required_fields = ["name", "address", "postalCode", "city", "standardPickup", "lat", "lon"]
    if not filename or not pickup or not all(field in pickup for field in required_fields):
        return jsonify({"error": "Puuttuvia kenttiä datasta."}), 400

    # Construct the file path of the Excel file.
    file_path = os.path.join(".secret", "ExcelFiles", f"{filename}.xlsx")
    if not os.path.exists(file_path):
        return jsonify({"error": "Excel-tiedostoa ei löydy."}), 404

    # Load the workbook and select the active worksheet.
    wb = load_workbook(file_path)
    ws = wb.active

    found = False
    # Iterate over route rows starting from row 3
    for row in ws.iter_rows(min_row=3, max_col=7):
        # Convert cell values into a list for comparison.
        values = [cell.value for cell in row]
        # Check if the current row matches the pickup data.
        if (
            str(values[0]) == str(pickup["name"]) and
            str(values[1]) == str(pickup["address"]) and
            str(values[2]) == str(pickup["postalCode"]) and
            str(values[3]) == str(pickup["city"]) and
            str(values[4]) == str(pickup["standardPickup"]) and
            str(values[5]) == str(pickup["lat"]) and
            str(values[6]) == str(pickup["lon"])
        ):
            # Delete the row if a match is found.
            row_index = row[0].row
            ws.delete_rows(row_index, 1)
            found = True
            break

    if not found:
        # Return an error if the pickup location was not found.
        return jsonify({"error": "Noutopaikkaa ei löytynyt Excel-tiedostosta."}), 404

    try:
        # Save the workbook after deletion.
        wb.save(file_path)
        return jsonify({"message": "Noutopaikka poistettu onnistuneesti Excel-tiedostosta."}), 200
    except Exception as e:
        # Return an error message if save fails.
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}), 500
    
"""
Updates the start and end times of an existing route in an Excel file.

:param file_name: str, name of the Excel file to update
:param startTime: str, updated start time
:param endTime: str, updated end time
:return: JSON response, success or failure message
"""
@excel_bp.route('/api/update_route_time', methods=['PUT'])
@cross_origin()
def update_route_time():
    # Parse JSON data from the request body
    data = request.get_json()

    # Retrieve file name and times from the request
    file_name = data.get("file_name")
    start_time = data.get("startTime")
    end_time = data.get("endTime")

    # Check that required fields are provided
    if not file_name or start_time is None or end_time is None:
        return jsonify({"error": "Puuttuvia kenttiä (file_name, startTime, endTime)."}), 400

    # Ensure the file has a valid .xlsx extension
    if not file_name.endswith(".xlsx"):
        file_name += ".xlsx"

    # Construct full path to the Excel file
    file_path = os.path.join(".secret", "ExcelFiles", file_name)

    # Check if the file exists
    if not os.path.exists(file_path):
        return jsonify({"error": "Tiedostoa ei löytynyt."}), 404

    try:
        # Load the Excel workbook and select active sheet
        wb = load_workbook(file_path)
        ws = wb.active

        # Update start and end times in cells N1 and N2
        ws["N1"] = start_time
        ws["N2"] = end_time

        # Save the updated Excel workbook
        wb.save(file_path)

        # Return success message
        return jsonify({"message": "Aika päivitetty onnistuneesti."}), 200

    except Exception as e:
        # Return error message if something goes wrong
        return jsonify({"error": True, "message": f"Ajan päivitys epäonnistui: {e}"}), 500

"""
Retrieves the original file names (C1) and save dates (D1) from each Excel file.

:return: JSON response, dictionary containing C1 and D1 values from all .xlsx files
"""
@excel_bp.route('/api/get_original_names', methods=['GET'])
@cross_origin()
def get_original_names():
    # Define the folder where Excel files are stored
    folder = os.path.join(".secret", "ExcelFiles")
    cells_data = {}
    
    # Check if the folder exists
    if not os.path.exists(folder):
        return jsonify({"error": "Kansiota ei löydy"}), 404

    # Iterate through all .xlsx files in the folder
    for file_name in os.listdir(folder):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder, file_name)
            try:
                # Load the Excel workbook and select the active sheet
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                # Read values from cells C1 and D1; return None if not available
                cell_C1 = ws["C1"].value
                cell_D1 = ws["D1"].value
                # Use the base filename (without extension) as the key
                base_name = os.path.splitext(file_name)[0]
                cells_data[base_name] = {"C1": cell_C1, "D1": cell_D1}
            except Exception as e:
                # Handle errors gracefully by returning them per file
                cells_data[file_name] = {"error": str(e)}

    # Return collected C1 and D1 values as JSON
    return jsonify(cells_data)

"""
Deletes all Excel files that were generated from same file

:param c1: str, the value to match in cell C1 of each Excel file
:return: JSON response listing deleted and failed files
"""
@excel_bp.route('/api/delete_by_group', methods=['DELETE'])
@cross_origin()
def delete_by_group():
    data = request.get_json()
    c1_value = data.get("c1")
    if not c1_value:
        return jsonify({"error": "alkuperäistä tiedostoa ei annettu"}), 400

    folder = os.path.join(".secret", "ExcelFiles")
    if not os.path.exists(folder):
        return jsonify({"error": "Tiedostokansiota ei löytynyt"}), 404

    deleted_files = []
    failed_files = []

    for filename in os.listdir(folder):
        if not filename.endswith(".xlsx"):
            continue

        path = os.path.join(folder, filename)
        try:
            # Lue C1-solun arvo
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            file_c1 = ws["C1"].value
            wb.close()

            # Vertaa siistittyinä ja pienennettyinä
            if str(file_c1).strip().lower() == str(c1_value).strip().lower():
                # Poista tiedosto
                os.remove(path)
                deleted_files.append(filename)
        except Exception as e:
            failed_files.append({"file": filename, "error": str(e)})

    return jsonify({
        "deleted": deleted_files,
        "failed": failed_files,
        "message": f"{len(deleted_files)} tiedostoa poistettu"
    }), 200

