from flask import Blueprint, jsonify, request
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from flask_cors import cross_origin

excel_bp = Blueprint('excel', __name__)

@excel_bp.route('/upload', methods=['POST'])
@cross_origin()
def upload_excel():
    data = request.get_json()
    route_name = data.get("routeName", "Uusi_reitti")
    file_name = data.get("fileName", "Tuntematon_tiedostonimi")
    excel_data = data.get("data", [])
    start_location = data.get("startLocation", {})
    end_location = data.get("endLocation", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Luetut paikat"

    # Tarkistetaan, onko tiedosto jo olemassa ja lisätään numerointi
    save_dir = os.path.join(".secret", "ExcelFiles")
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)

    original_file_name = os.path.splitext(file_name)[0]
    file_path = os.path.join(save_dir, f"{route_name} ({original_file_name}).xlsx")
    file_counter = 1
    if os.path.exists(file_path):
        while os.path.exists(file_path):
            file_path = os.path.join(save_dir, f"{route_name} {file_counter} ({original_file_name}).xlsx")
            file_counter += 1
    
    
    final_file_name = os.path.splitext(os.path.basename(file_path))[0] + " (" + original_file_name + ")"
    ws["A1"] = "Reitin nimi:"
    ws["B1"] = final_file_name

    header = ["Nimi", "Osoite", "Postinumero", "Kaupunki", "Vakionouto", "Lat", "Lon"]
    ws.append(header)

    # Kirjoitetaan aloituspaikan tiedot solusta I1 alkaen, jos tiedot ovat saatavilla
    if start_location:
        ws["I1"] = start_location.get("name")
        ws["J1"] = start_location.get("address")
        ws["K1"] = start_location.get("postalCode")
        ws["L1"] = start_location.get("city")
        ws["M1"] = "yes"
        ws["N1"] = start_location.get("departureTime")
        ws["O1"] = start_location.get("lat")
        ws["P1"] = start_location.get("lon")

    if end_location:
        ws["I2"] = end_location.get("name")
        ws["J2"] = end_location.get("address")
        ws["K2"] = end_location.get("postalCode")
        ws["L2"] = end_location.get("city")
        ws["M2"] = "yes"
        ws["N2"] = end_location.get("endTime")
        ws["O2"] = end_location.get("lat")
        ws["P2"] = end_location.get("lon")

    # Lisätään muut reitin pisteet
    row_num = 3
    for item in excel_data:
        ws.cell(row=row_num, column=1, value=item.get("name"))
        ws.cell(row=row_num, column=2, value=item.get("address"))
        ws.cell(row=row_num, column=3, value=item.get("postalCode"))
        ws.cell(row=row_num, column=4, value=item.get("city"))
        ws.cell(row=row_num, column=5, value=item.get("standardPickup"))
        ws.cell(row=row_num, column=6, value=item.get("lat"))
        ws.cell(row=row_num, column=7, value=item.get("lon"))
        row_num += 1

    try:
        wb.save(file_path)
        return jsonify({"message": "Tiedosto tallennettu onnistuneesti", "file_path": file_path}), 200
    except Exception as e:
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}), 500
    
@excel_bp.route('/api/get_excel_files', methods=['GET'])
@cross_origin()
def get_excel_jsons():
    folder = os.path.join(".secret", "ExcelFiles")
    excel_jsons = {}
    
    if os.path.exists(folder):
        for file_name in os.listdir(folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(folder, file_name)
                try:
                    # Lue tiedoston nimi ilman laajennusta
                    base_name = os.path.splitext(file_name)[0]
                    excel_jsons[base_name] = {"file_name": file_name, "path": file_path}
                except Exception as e:
                    print(f"Virhe tiedostoa {file_name} lukiessa: {e}")
                    base_name = os.path.splitext(file_name)[0]
                    excel_jsons[base_name] = {"error": str(e)}
    else:
        return jsonify({"error": "Kansiota ei löydy"}), 404
    
    return jsonify(excel_jsons)

@excel_bp.route('/api/delete_excel', methods=['DELETE'])
@cross_origin()
def delete_excel():
    data = request.get_json()
    file_name = data.get("file_name")
    if not file_name:
        return jsonify({"error": "file_name not provided"}), 400

    if not file_name.endswith(".xlsx"):
        file_name_full = file_name + ".xlsx"
    else:
        file_name_full = file_name

    folder = os.path.join(".secret", "ExcelFiles")
    file_path = os.path.join(folder, file_name_full)
    
    if not os.path.exists(file_path):
        return jsonify({"error": "Tiedostoa ei löytynyt"}), 404

    try:
        os.remove(file_path)
        return jsonify({"message": "Tiedosto poistettu onnistuneesti", "file_name": file_name_full}), 200
    except Exception as e:
        return jsonify({"error": True, "message": f"Tiedoston poisto epäonnistui: {e}"}), 500
    
@excel_bp.route('/api/get_route', methods=['GET'])
@cross_origin()
def get_excel_routes():
    folder = os.path.join(".secret", "ExcelFiles")
    excel_routes = {}
    
    if not os.path.exists(folder):
        return jsonify({"error": "Kansiota ei löydy"}), 404
    
    for file_name in os.listdir(folder):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder, file_name)
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active

                # Lue reitin nimi solusta B1
                route_name = ws["B1"].value

                # Lue aloituspaikan tiedot solusta I1 - P1
                startPlace = {
                    "name": ws["I1"].value,
                    "address": ws["J1"].value,
                    "postalCode": ws["K1"].value,
                    "city": ws["L1"].value,
                    "standardPickup": ws["M1"].value,
                    "lat": ws["O1"].value,
                    "lon": ws["P1"].value,
                }
                
                # Lue loppupisteen tiedot solusta I2 - P2
                endPlace = {
                    "name": ws["I2"].value,
                    "address": ws["J2"].value,
                    "postalCode": ws["K2"].value,
                    "city": ws["L2"].value,
                    "standardPickup": ws["M2"].value,
                    "lat": ws["O2"].value,
                    "lon": ws["P2"].value,
                }
                
                # Aloitus- ja lopetusajat solusta N1 ja N2
                startTime = ws["N1"].value
                endTime = ws["N2"].value

                # Reitit alkavat riviltä 3.
                routes = []
                # Oletuksena solut A-G sisältävät reittien tiedot.
                for row in ws.iter_rows(min_row=3, max_col=7):
                    # Tarkistetaan, ettei rivi ole kokonaan tyhjä
                    if not any(cell.value for cell in row):
                        continue
                    route_entry = {
                        "name": row[0].value,        # A-solu
                        "address": row[1].value,     # B-solu
                        "postalCode": row[2].value,  # C-solu
                        "city": row[3].value,        # D-solu
                        "standardPickup": row[4].value,  # E-solu
                        "lat": row[5].value,         # F-solu
                        "lon": row[6].value,         # G-solu
                    }
                    routes.append(route_entry)
                
                route_json = {
                    "name": route_name,
                    "startPlace": startPlace,
                    "endPlace": endPlace,
                    "startTime": startTime,
                    "endTime": endTime,
                    "routes": routes,
                }
                excel_routes[file_name] = route_json
                
            except Exception as e:
                excel_routes[file_name] = {"error": str(e)}
    
    return jsonify(excel_routes)

@excel_bp.route('/api/append_to_excel', methods=['POST'])
@cross_origin()
def append_to_excel():
    data = request.get_json()
    filename = data.get("filename")
    pickup = data.get("data")

    required_fields = ["name", "address", "postalCode", "city", "standardPickup", "lat", "lon"]
    if not filename or not all(field in pickup for field in required_fields):
        return jsonify({"error": "Puuttuvia kenttiä datasta."}), 400

    file_path = os.path.join(".secret", "ExcelFiles", f"{filename}.xlsx")
    if not os.path.exists(file_path):
        return jsonify({"error": "Excel-tiedostoa ei löydy."}), 404

    wb = load_workbook(file_path)
    ws = wb.active

    ws.append([
        pickup["name"],
        pickup["address"],
        pickup["postalCode"],
        pickup["city"],
        pickup["standardPickup"],
        pickup["lat"],
        pickup["lon"]
    ])

    try:
        wb.save(file_path)
        return jsonify({"message": "Paikka lisätty onnistuneesti."}), 200
    except Exception as e:
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}),

@excel_bp.route('/api/remove_from_excel', methods=['POST'])
@cross_origin()
def remove_from_excel():
    data = request.get_json()
    filename = data.get("filename")
    pickup = data.get("data")

    required_fields = ["name", "address", "postalCode", "city", "standardPickup", "lat", "lon"]
    if not filename or not pickup or not all(field in pickup for field in required_fields):
        return jsonify({"error": "Puuttuvia kenttiä datasta."}), 400

    file_path = os.path.join(".secret", "ExcelFiles", f"{filename}.xlsx")
    if not os.path.exists(file_path):
        return jsonify({"error": "Excel-tiedostoa ei löydy."}), 404

    wb = load_workbook(file_path)
    ws = wb.active

    found = False
    for row in ws.iter_rows(min_row=3, max_col=7):
        values = [cell.value for cell in row]
        if (
            str(values[0]) == str(pickup["name"]) and
            str(values[1]) == str(pickup["address"]) and
            str(values[2]) == str(pickup["postalCode"]) and
            str(values[3]) == str(pickup["city"]) and
            str(values[4]) == str(pickup["standardPickup"]) and
            str(values[5]) == str(pickup["lat"]) and
            str(values[6]) == str(pickup["lon"])
        ):
            row_index = row[0].row
            ws.delete_rows(row_index, 1)
            found = True
            break

    if not found:
        return jsonify({"error": "Noutopaikkaa ei löytynyt Excel-tiedostosta."}), 404

    try:
        wb.save(file_path)
        return jsonify({"message": "Noutopaikka poistettu onnistuneesti Excel-tiedostosta."}), 200
    except Exception as e:
        return jsonify({"error": True, "message": f"Tiedoston tallennus epäonnistui: {e}"}), 500